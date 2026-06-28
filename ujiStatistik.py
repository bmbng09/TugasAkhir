import pandas as pd
import numpy as np
from scipy import stats
from itertools import combinations
import matplotlib.pyplot as plt
import matplotlib
matplotlib.rcParams['font.family'] = 'DejaVu Sans'
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

print("=" * 65)
print("ANALISIS STATISTIK - FRIEDMAN TEST + WILCOXON SIGNED-RANK TEST")
print("=" * 65)

#  1. Load & clean data 
print("\n[1/6] Membaca data skor final...")
df = pd.read_excel("Data_LLM_AS_A_JUDGE.xlsx",
                   sheet_name="Skor Final", header=1)
df.columns = ['ID_Q','Pertanyaan','Sub_Bidang','Teknik',
              'RP','AI','KR','KC','KO','Rata_rata']
df = df.dropna(subset=['Teknik','RP']).reset_index(drop=True)
for col in ['RP','AI','KR','KC','KO','Rata_rata']:
    df[col] = pd.to_numeric(df[col], errors='coerce')
df = df[df['Teknik'].isin(['Zero-Shot','Few-Shot','CoT','Role','Hybrid'])]

TEKNIK = ['Zero-Shot','Few-Shot','CoT','Role','Hybrid']
TEKNIK_SHORT = ['ZS','FS','CoT','Role','Hyb']
DIMS = ['RP','AI','KR','KC','KO','Rata_rata']
DIM_LABELS = ['RP','AI','KR','KC','KO','Rata-rata']

print(f"  Total baris: {len(df)}")
print(f"  Teknik: {df['Teknik'].unique().tolist()}")
print(f"  Sub-bidang: {df['Sub_Bidang'].dropna().unique().tolist()}")

#  2. Build matriks 50x5 per dimensi 
print("\n[2/6] Membangun matriks skor...")
questions = sorted(df['ID_Q'].unique())

def build_matrix(dim):
    mat = pd.DataFrame(index=questions, columns=TEKNIK, dtype=float)
    for q in questions:
        for t in TEKNIK:
            val = df[(df['ID_Q']==q) & (df['Teknik']==t)][dim].values
            mat.loc[q, t] = val[0] if len(val) > 0 else np.nan
    return mat.dropna()

matrices = {dim: build_matrix(dim) for dim in DIMS}
print(f"  Matriks bersih: {matrices['Rata_rata'].shape} per dimensi")

#  3. Friedman Test — Agregat 
print("\n[3/6] Menjalankan Friedman Test...")
print("\n  === FRIEDMAN TEST AGREGAT ===")

friedman_results = {}
for dim, label in zip(DIMS, DIM_LABELS):
    mat = matrices[dim]
    groups = [mat[t].values for t in TEKNIK]
    stat, p = stats.friedmanchisquare(*groups)
    n = len(mat)
    df_val = len(TEKNIK) - 1
    friedman_results[dim] = {'stat': stat, 'p': p, 'n': n, 'df': df_val}

    # Mean ranks
    ranked = mat.rank(axis=1)
    mean_ranks = ranked.mean()

    sig = "SIGNIFIKAN" if p < 0.05 else "Tidak signifikan"
    print(f"  {label:12}: χ²F={stat:.4f}, p={p:.6f} {sig}")
    if dim == 'Rata_rata':
        print(f"    Mean ranks: {dict(zip(TEKNIK_SHORT, [f'{mean_ranks[t]:.4f}' for t in TEKNIK]))}")

#  4. Wilcoxon Post-Hoc 
print("\n[4/6] Menjalankan Wilcoxon Signed-Rank Test...")
pairs = list(combinations(TEKNIK, 2))
n_pairs = len(pairs)
alpha_bonferroni = 0.05 / n_pairs

print(f"\n  Jumlah pasangan: {n_pairs}")
print(f"  α Bonferroni: 0.05 / {n_pairs} = {alpha_bonferroni:.4f}")
print(f"\n  === WILCOXON PAIRWISE (Rata-rata 5 dimensi) ===")

mat_overall = matrices['Rata_rata']
wilcoxon_results = []

for (t1, t2) in pairs:
    x = mat_overall[t1].values
    y = mat_overall[t2].values
    try:
        W, p = stats.wilcoxon(x, y, alternative='two-sided')
    except Exception as e:
        W, p = np.nan, np.nan

    mean_diff = np.mean(x) - np.mean(y)
    sig_bonf = p < alpha_bonferroni if not np.isnan(p) else False
    sig_alpha = p < 0.05 if not np.isnan(p) else False

    # Effect size r = Z / sqrt(n)
    if not np.isnan(p) and p > 0:
        z = abs(stats.norm.ppf(p / 2))
        r = z / np.sqrt(len(x))
    else:
        r = np.nan

    mag = "besar" if not np.isnan(r) and r >= 0.5 else \
          "sedang" if not np.isnan(r) and r >= 0.3 else "kecil"

    result = {
        'T1': t1, 'T2': t2, 'W': W, 'p': p,
        'sig_bonf': sig_bonf, 'sig_alpha': sig_alpha,
        'mean_diff': mean_diff, 'r': r, 'mag': mag
    }
    wilcoxon_results.append(result)

    mark = "SIGNIFIKAN" if sig_bonf else ("~ mendekati" if sig_alpha else "")
    print(f"  {t1:12} vs {t2:12}: W={W:.1f}, p={p:.6f} {mark}")
    print(f"    mean_diff={mean_diff:+.4f}, r={r:.4f} ({mag})")

#  5. Friedman per sub-bidang 
print("\n[5/6] Friedman Test per sub-bidang...")
sub_results = []
for sub in sorted(df['Sub_Bidang'].dropna().unique()):
    sub_df = df[df['Sub_Bidang']==sub]
    sub_q = sorted(sub_df['ID_Q'].unique())
    mat_sub = pd.DataFrame(index=sub_q, columns=TEKNIK, dtype=float)
    for q in sub_q:
        for t in TEKNIK:
            val = sub_df[(sub_df['ID_Q']==q) & (sub_df['Teknik']==t)]['Rata_rata'].values
            mat_sub.loc[q, t] = val[0] if len(val) > 0 else np.nan
    mat_sub = mat_sub.dropna()
    n_sub = len(mat_sub)
    if n_sub >= 3:
        grps = [mat_sub[t].values for t in TEKNIK]
        s, p_sub = stats.friedmanchisquare(*grps)
        sig = p_sub < 0.05
        sub_results.append({'sub': sub, 'n': n_sub, 'chi2': s, 'p': p_sub, 'sig': sig})
        mark = "SIGNIFIKAN" if sig else "Tidak signifikan"
        print(f"  {sub[:40]:40} n={n_sub}, χ²={s:.3f}, p={p_sub:.4f} {mark}")

#  6. Save results to Excel 
print("\n[6/6] Menyimpan hasil ke Excel...")

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()

#  Sheet 1: Friedman Agregat 
ws1 = wb.active
ws1.title = "Friedman Agregat"
ws1.sheet_view.showGridLines = False

col_widths = {'A':22,'B':12,'C':8,'D':12,'E':14,'F':24}
for col, w in col_widths.items(): ws1.column_dimensions[col].width = w

ws1.row_dimensions[1].height = 26
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "FRIEDMAN TEST — Agregat dan Per Dimensi (n=50, df=4, α=0,05)"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("1F4E79")
c.alignment = align()

headers = ["Dimensi","χ²F","df","p-value","Keputusan","Keterangan"]
ws1.row_dimensions[2].height = 26
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill("2E75B6")
    c.alignment = align()
    c.border = border_thin()

for i, (dim, label) in enumerate(zip(DIMS, DIM_LABELS)):
    row = i + 3
    ws1.row_dimensions[row].height = 22
    r = friedman_results[dim]
    bg = "EBF3FB" if dim == "Rata_rata" else ("F5F9FF" if i%2==0 else "FFFFFF")
    sig = r['p'] < 0.05
    sig_text = "Tolak H₀" if sig else "Gagal tolak H₀"
    ket = "Perbedaan signifikan" if sig else "Tidak ada perbedaan signifikan"

    for col, val in enumerate([label, f"{r['stat']:.4f}", str(r['df']),
                                f"{r['p']:.6f}", sig_text, ket], 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = font(size=10, bold=(dim=="Rata_rata"),
                      color="375623" if (sig and col in [4,5]) else "000000")
        c.fill = fill(bg)
        c.alignment = align(h="center" if col > 1 else "left")
        c.border = border_thin()

#  Sheet 2: Mean Ranks 
ws2 = wb.create_sheet("Mean Ranks")
ws2.sheet_view.showGridLines = False
for col in range(1, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 14

ws2.row_dimensions[1].height = 26
ws2.merge_cells(f"A1:{get_column_letter(len(TEKNIK)+1)}1")
c = ws2["A1"]
c.value = "MEAN RANKS FRIEDMAN TEST PER DIMENSI"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("1F4E79")
c.alignment = align()

ws2.row_dimensions[2].height = 26
ws2.cell(row=2, column=1, value="Dimensi").font = font(bold=True, color="FFFFFF", size=10)
ws2.cell(row=2, column=1).fill = fill("1F4E79")
ws2.cell(row=2, column=1).alignment = align()
ws2.cell(row=2, column=1).border = border_thin()

for j, t in enumerate(TEKNIK):
    c = ws2.cell(row=2, column=j+2, value=t)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill("2E75B6")
    c.alignment = align()
    c.border = border_thin()

for i, (dim, label) in enumerate(zip(DIMS, DIM_LABELS)):
    row = i + 3
    ws2.row_dimensions[row].height = 22
    bg = "EBF3FB" if dim == "Rata_rata" else ("F5F9FF" if i%2==0 else "FFFFFF")
    mat = matrices[dim]
    ranked = mat.rank(axis=1)
    mean_ranks = ranked.mean()
    max_rank = mean_ranks.max()

    c = ws2.cell(row=row, column=1, value=label)
    c.font = font(bold=(dim=="Rata_rata"), size=10)
    c.fill = fill(bg)
    c.alignment = align(h="left")
    c.border = border_thin()

    for j, t in enumerate(TEKNIK):
        rj = mean_ranks[t]
        is_max = rj == max_rank
        c = ws2.cell(row=row, column=j+2, value=round(rj, 4))
        c.number_format = "0.0000"
        c.font = font(bold=is_max, size=10, color="375623" if is_max else "000000")
        c.fill = fill("E2F0D9" if is_max else bg)
        c.alignment = align()
        c.border = border_thin()

#  Sheet 3: Wilcoxon 
ws3 = wb.create_sheet("Wilcoxon Post-Hoc")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 10
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 14

ws3.row_dimensions[1].height = 26
ws3.merge_cells("A1:I1")
c = ws3["A1"]
c.value = f"WILCOXON SIGNED-RANK TEST — 10 Pasangan | α_Bonferroni = {alpha_bonferroni:.4f}"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("1F4E79")
c.alignment = align()

headers3 = ["Teknik A","Teknik B","W","p-value","p < 0,005\n(Bonferroni)","p < 0,05","Selisih Mean","Effect r","Magnitud"]
ws3.row_dimensions[2].height = 32
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill("2E75B6")
    c.alignment = align()
    c.border = border_thin()

for i, r in enumerate(wilcoxon_results):
    row = i + 3
    ws3.row_dimensions[row].height = 22
    bg = "F5F9FF" if i%2==0 else "FFFFFF"
    if r['sig_bonf']: bg = "E2F0D9"
    elif r['sig_alpha']: bg = "FFF2CC"

    vals = [r['T1'], r['T2'],
            f"{r['W']:.1f}" if not np.isnan(r['W']) else "—",
            f"{r['p']:.6f}" if not np.isnan(r['p']) else "—",
            "SIGNIFIKAN" if r['sig_bonf'] else "Tidak signifikan",
            "SIGNIFIKAN" if r['sig_alpha'] else "Tidak signifikan",
            f"{r['mean_diff']:+.4f}",
            f"{r['r']:.4f}" if not np.isnan(r['r']) else "—",
            r['mag']]

    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        color = "375623" if r['sig_bonf'] else "7F4F00" if r['sig_alpha'] else "000000"
        c.font = font(size=10, bold=(col in [4,5] and r['sig_alpha']), color=color)
        c.fill = fill(bg)
        c.alignment = align(h="center" if col > 2 else "left")
        c.border = border_thin()

#  Sheet 4: Matriks p-value 
ws4 = wb.create_sheet("Matriks p-value")
ws4.sheet_view.showGridLines = False
for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 15

ws4.row_dimensions[1].height = 26
ws4.merge_cells(f"A1:{get_column_letter(len(TEKNIK)+1)}1")
c = ws4["A1"]
c.value = "MATRIKS p-VALUE WILCOXON (Hijau=p<0,005 | Kuning=p<0,05)"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("1F4E79")
c.alignment = align()

ws4.cell(row=2, column=1, value="Teknik").font = font(bold=True, color="FFFFFF", size=10)
ws4.cell(row=2, column=1).fill = fill("1F4E79")
ws4.cell(row=2, column=1).alignment = align()
ws4.cell(row=2, column=1).border = border_thin()
for j, t in enumerate(TEKNIK):
    c = ws4.cell(row=2, column=j+2, value=t)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill("2E75B6")
    c.alignment = align()
    c.border = border_thin()

p_dict = {}
for r in wilcoxon_results:
    p_dict[(r['T1'], r['T2'])] = r['p']
    p_dict[(r['T2'], r['T1'])] = r['p']

for i, t1 in enumerate(TEKNIK):
    row = i + 3
    ws4.row_dimensions[row].height = 22
    bg = "F5F9FF" if i%2==0 else "FFFFFF"
    c = ws4.cell(row=row, column=1, value=t1)
    c.font = font(bold=True, size=10)
    c.fill = fill(bg)
    c.alignment = align(h="left")
    c.border = border_thin()
    for j, t2 in enumerate(TEKNIK):
        c = ws4.cell(row=row, column=j+2)
        if t1 == t2:
            c.value = "—"
            c.fill = fill("EEEEEE")
            c.font = font(color="888888")
        else:
            p_val = p_dict.get((t1, t2), np.nan)
            if not np.isnan(p_val):
                c.value = round(p_val, 6)
                c.number_format = "0.0000"
                if p_val < 0.005:
                    c.fill = fill("E2F0D9")
                    c.font = font(bold=True, color="375623", size=10)
                elif p_val < 0.05:
                    c.fill = fill("FFF2CC")
                    c.font = font(bold=True, color="7F4F00", size=10)
                else:
                    c.fill = fill(bg)
                    c.font = font(size=10)
        c.alignment = align()
        c.border = border_thin()

#  Sheet 5: Per Sub-Bidang 
ws5 = wb.create_sheet("Friedman Per Sub-Bidang")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 36
ws5.column_dimensions['B'].width = 8
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 14
ws5.column_dimensions['F'].width = 22

ws5.row_dimensions[1].height = 26
ws5.merge_cells("A1:F1")
c = ws5["A1"]
c.value = "FRIEDMAN TEST PER SUB-BIDANG (Analisis Eksploratori, α=0,05, df=4)"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("1F4E79")
c.alignment = align()

headers5 = ["Sub-Bidang","n","χ²F","p-value","Signifikan","Interpretasi"]
ws5.row_dimensions[2].height = 26
for col, h in enumerate(headers5, 1):
    c = ws5.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill("2E75B6")
    c.alignment = align()
    c.border = border_thin()

for i, r in enumerate(sub_results):
    row = i + 3
    ws5.row_dimensions[row].height = 22
    bg = "E2F0D9" if r['sig'] else ("F5F9FF" if i%2==0 else "FFFFFF")
    sig_text = "Ya" if r['sig'] else "Tidak"
    ket = "Ada perbedaan signifikan antar teknik" if r['sig'] else "Tidak ada perbedaan signifikan"
    for col, val in enumerate([r['sub'], r['n'], round(r['chi2'],3),
                                round(r['p'],6), sig_text, ket], 1):
        c = ws5.cell(row=row, column=col, value=val)
        c.font = font(size=10, bold=r['sig'], color="375623" if r['sig'] else "000000")
        c.fill = fill(bg)
        c.alignment = align(h="center" if col in [2,3,4,5] else "left")
        c.border = border_thin()

#  Save 
wb.save("hasil_statistik.xlsx")

#  Visualisasi: Mean Ranks 
mat_overall = matrices['Rata_rata']
ranked_overall = mat_overall.rank(axis=1)
mean_ranks_overall = ranked_overall.mean()

fig, axes = plt.subplots(1, 2, figsize=(14, 5))

# Plot 1: Mean ranks bar chart
colors = ['#1F4E79','#7F4F00','#1E5628','#5B0070','#7F0000']
bars = axes[0].bar(TEKNIK, [mean_ranks_overall[t] for t in TEKNIK], color=colors, edgecolor='white', linewidth=1.5)
axes[0].axhline(y=3.0, color='gray', linestyle='--', alpha=0.5, label='Rata-rata teoritis (3.0)')
axes[0].set_title('Mean Rank Friedman Test\n(semakin tinggi = semakin baik)', fontweight='bold')
axes[0].set_ylabel('Mean Rank')
axes[0].set_ylim(0, 4.5)
for bar, val in zip(bars, [mean_ranks_overall[t] for t in TEKNIK]):
    axes[0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
                f'{val:.3f}', ha='center', va='bottom', fontweight='bold', fontsize=10)
axes[0].tick_params(axis='x', rotation=15)
axes[0].legend()

# Plot 2: p-value heatmap
p_matrix = np.ones((5, 5))
for r in wilcoxon_results:
    i = TEKNIK.index(r['T1'])
    j = TEKNIK.index(r['T2'])
    p_matrix[i][j] = r['p']
    p_matrix[j][i] = r['p']
np.fill_diagonal(p_matrix, np.nan)

df_p = pd.DataFrame(p_matrix, index=TEKNIK_SHORT, columns=TEKNIK_SHORT)
mask = np.eye(5, dtype=bool)
sns.heatmap(df_p, annot=True, fmt='.4f', cmap='RdYlGn_r',
            mask=mask, ax=axes[1], vmin=0, vmax=0.1,
            linewidths=0.5, cbar_kws={'shrink': 0.8},
            annot_kws={'size': 9})
axes[1].set_title('Matriks p-value Wilcoxon\n(hijau=rendah/signifikan, merah=tinggi)',
                  fontweight='bold')

# Add significance markers
for i in range(5):
    for j in range(5):
        if i != j and p_matrix[i][j] < 0.05:
            axes[1].add_patch(plt.Rectangle((j, i), 1, 1, fill=False,
                                           edgecolor='blue', lw=2))

plt.tight_layout()
plt.savefig('plot_statistik.png', dpi=150, bbox_inches='tight')
plt.close()

print("\nFile output:")
print("  hasil_statistik.xlsx — 5 sheet hasil lengkap")
print("  plot_statistik.png   — visualisasi mean ranks + matriks p-value")

#  Summary 
print("\n" + "=" * 65)
print("RINGKASAN HASIL")
print("=" * 65)
r_overall = friedman_results['Rata_rata']
print(f"\nFRIEDMAN TEST (Agregat):")
print(f"  χ²F = {r_overall['stat']:.4f}, df = {r_overall['df']}, p = {r_overall['p']:.6f}")
print(f"  {'H₀ DITOLAK' if r_overall['p'] < 0.05 else 'H₀ GAGAL DITOLAK'}")

print(f"\nRANKING TEKNIK (berdasarkan mean rank Friedman):")
mean_ranks_overall_sorted = sorted(TEKNIK, key=lambda t: mean_ranks_overall[t], reverse=True)
for rank, t in enumerate(mean_ranks_overall_sorted, 1):
    print(f"  {rank}. {t:15}: Rj={mean_ranks_overall[t]:.4f}, mean skor={mat_overall[t].mean():.4f}")

print(f"\nWILCOXON POST-HOC:")
sig_bonf = [r for r in wilcoxon_results if r['sig_bonf']]
sig_alpha = [r for r in wilcoxon_results if r['sig_alpha'] and not r['sig_bonf']]
print(f"  Signifikan (p < {alpha_bonferroni:.4f}): {len(sig_bonf)} pasangan")
print(f"  Mendekati signifikan (p < 0,05): {len(sig_alpha)} pasangan")
for r in sorted(wilcoxon_results, key=lambda x: x['p'])[:3]:
    print(f"  {r['T1']:12} vs {r['T2']:12}: p={r['p']:.6f}, r={r['r']:.3f}")

print("\nSelesai!")
