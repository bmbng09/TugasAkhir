import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

#  1. Load data 
print("=" * 70)
print("COSINE SIMILARITY + BERTSCORE — Konsistensi Antar Run (v2 fixed)")
print("=" * 70)

print("\n[1/6] Membaca Data_Eksperimen.xlsx...")
src = pd.read_excel("Data Eksperimen\\Data_Eksperimen.xlsx", sheet_name="Log Eksperimen", header=1)
src.columns = ['ID_Sesi','ID_Pertanyaan','Pertanyaan','Sub_Bidang',
               'Teknik','Run','Prompt','Output','Status']
src['Run'] = pd.to_numeric(src['Run'], errors='coerce').fillna(0).astype(int)
src['Output'] = src['Output'].astype(str)
src = src[src['Run'].isin([1,2,3,4,5])].reset_index(drop=True)

TEKNIK = ["Zero-Shot", "Few-Shot", "CoT", "Role", "Hybrid"]
pertanyaan_ids = sorted(src['ID_Pertanyaan'].unique())
sub_bidang_map = src.groupby('ID_Pertanyaan')['Sub_Bidang'].first().to_dict()
run_pairs = list(combinations([1,2,3,4,5], 2))  # 10 pasangan

print(f"  Total sesi valid  : {len(src)}")
print(f"  Pertanyaan        : {len(pertanyaan_ids)}")
print(f"  Teknik            : {len(TEKNIK)}")
print(f"  Pasangan run      : {len(run_pairs)} (C(5,2))")
print(f"  Total unit        : {len(pertanyaan_ids)*len(TEKNIK)}")
print(f"  Total perhitungan : {len(pertanyaan_ids)*len(TEKNIK)*len(run_pairs)} per metode")

# Sanity check: run 1 vs run 2 harus berbeda
pid0 = pertanyaan_ids[0]
r1_sample = src[(src['ID_Pertanyaan']==pid0)&(src['Teknik']=='Zero-Shot')&(src['Run']==1)]['Output'].values[0]
r2_sample = src[(src['ID_Pertanyaan']==pid0)&(src['Teknik']=='Zero-Shot')&(src['Run']==2)]['Output'].values[0]
print(f"\n  Sanity check ({pid0}, Zero-Shot):")
print(f"  Run 1 ({len(r1_sample)} chars): {r1_sample[:80]!r}")
print(f"  Run 2 ({len(r2_sample)} chars): {r2_sample[:80]!r}")
print(f"  Identik? {r1_sample==r2_sample} (harus False)")

#  2. Buat lookup dict output 
print("\n[2/6] Membangun lookup dictionary output...")
output_lookup = {}
for _, row in src.iterrows():
    key = (row['ID_Pertanyaan'], row['Teknik'], int(row['Run']))
    output_lookup[key] = str(row['Output'])
print(f"  Total entri lookup: {len(output_lookup)}")

#  3. Load model 
print("\n[3/6] Memuat Sentence Transformer...")
st_model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
print("  Model: paraphrase-multilingual-MiniLM-L12-v2 (384 dimensi)")


# FUNGSI COSINE 

def cosine_manual(vec_a: np.ndarray, vec_b: np.ndarray) -> float:
    """
    Formula Cosine Similarity:

              A . B
    cos(θ) = ----------
             ||A|| x ||B||

    A . B = Σ(i=1..384) A_i x B_i    (dot product)
    ||A|| = sqrt(Σ A_i²)              (norma Euclidean)
    ||B|| = sqrt(Σ B_i²)
    """
    dot = np.sum(vec_a * vec_b)
    na  = np.sqrt(np.sum(vec_a ** 2))
    nb  = np.sqrt(np.sum(vec_b ** 2))
    if na == 0 or nb == 0:
        return 0.0
    return float(dot / (na * nb))


#  4. Hitung Cosine Similarity antar run 
print("\n[4/6] Menghitung Cosine Similarity antar run (2.500 perhitungan)...")

# Verifikasi formula manual dulu
from sklearn.metrics.pairwise import cosine_similarity as sklearn_cs
emb_r1 = st_model.encode([r1_sample])[0]
emb_r2 = st_model.encode([r2_sample])[0]
manual_v  = cosine_manual(emb_r1, emb_r2)
sklearn_v = sklearn_cs([emb_r1], [emb_r2])[0][0]
print(f"\n  Verifikasi formula manual vs sklearn ({pid0}, ZS, run1-run2):")
print(f"    Manual  = {manual_v:.10f}")
print(f"    Sklearn = {sklearn_v:.10f}")
print(f"    Selisih = {abs(manual_v-sklearn_v):.2e}  (harus ~nol)")
print(f"    Nilai   = {manual_v:.4f} (TIDAK boleh 1.0 karena teks berbeda)\n")

cosine_detail_rows = []
cosine_unit_rows   = []
done = 0
total = len(pertanyaan_ids) * len(TEKNIK) * len(run_pairs)

for pid in pertanyaan_ids:
    sub = sub_bidang_map[pid]
    for teknik in TEKNIK:
        pair_scores = []

        for (r1, r2) in run_pairs:
            text_r1 = output_lookup.get((pid, teknik, r1), "")
            text_r2 = output_lookup.get((pid, teknik, r2), "")

            if not text_r1 or not text_r2:
                print(f"  SKIP: {pid} {teknik} run {r1}/{r2} tidak ditemukan")
                continue

            # Encode masing-masing secara individual
            emb_a = st_model.encode([text_r1])[0]
            emb_b = st_model.encode([text_r2])[0]
            val   = cosine_manual(emb_a, emb_b)

            pair_scores.append(val)
            cosine_detail_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'Run_A': r1, 'Run_B': r2, 'Cosine': round(val, 6),
            })

            done += 1
            if done % 100 == 0:
                print(f"  Progress cosine: {done}/{total} ({done/total*100:.0f}%)")

        if pair_scores:
            cosine_unit_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'Cosine_Konsistensi': round(np.mean(pair_scores), 6),
                'Cosine_Min':  round(min(pair_scores), 6),
                'Cosine_Max':  round(max(pair_scores), 6),
                'Cosine_Std':  round(np.std(pair_scores), 6),
            })

df_cosine_detail = pd.DataFrame(cosine_detail_rows)
df_cosine_unit   = pd.DataFrame(cosine_unit_rows)

cosine_per_teknik = (df_cosine_unit.groupby('Teknik')['Cosine_Konsistensi']
                     .agg(['mean','min','max','std']).round(4))
cosine_per_teknik.columns = ['Mean','Min','Max','Std']
cosine_per_teknik = cosine_per_teknik.reindex(TEKNIK)

print("\n  Skor Konsistensi Cosine per Teknik (rata-rata 50 pertanyaan):")
print(f"  {'Teknik':15} {'Mean':8} {'Min':8} {'Max':8} {'Std':8}")
for t in TEKNIK:
    r = cosine_per_teknik.loc[t]
    print(f"  {t:15} {r['Mean']:.4f}   {r['Min']:.4f}   {r['Max']:.4f}   {r['Std']:.4f}")

#  5. BERTScore antar run (apple-to-apple) 
print("\n[5/6] Menghitung BERTScore antar run (2.500 perhitungan, apple-to-apple)...")
bs_available = False
try:
    from bert_score import score as bs_score
    bs_available = True

    bs_detail_rows = []
    bs_unit_rows   = []
    done_bs = 0

    for pid in pertanyaan_ids:
        sub = sub_bidang_map[pid]
        for teknik in TEKNIK:
            pair_scores = []

            for (r1, r2) in run_pairs:
                text_r1 = output_lookup.get((pid, teknik, r1), "")
                text_r2 = output_lookup.get((pid, teknik, r2), "")

                if not text_r1 or not text_r2:
                    continue

                # Dua arah → simetris
                _, _, F1_ab = bs_score([text_r1], [text_r2], lang="id", verbose=False)
                _, _, F1_ba = bs_score([text_r2], [text_r1], lang="id", verbose=False)
                val_sym = (F1_ab.item() + F1_ba.item()) / 2

                pair_scores.append(val_sym)
                bs_detail_rows.append({
                    'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                    'Run_A': r1, 'Run_B': r2,
                    'BS_AB': round(F1_ab.item(), 6),
                    'BS_BA': round(F1_ba.item(), 6),
                    'BS_Sym': round(val_sym, 6),
                })

                done_bs += 1
                if done_bs % 100 == 0:
                    print(f"  Progress BERTScore: {done_bs}/{total} ({done_bs/total*100:.0f}%)")

            if pair_scores:
                bs_unit_rows.append({
                    'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                    'BS_Konsistensi': round(np.mean(pair_scores), 6),
                    'BS_Min': round(min(pair_scores), 6),
                    'BS_Max': round(max(pair_scores), 6),
                    'BS_Std': round(np.std(pair_scores), 6),
                })

    df_bs_detail = pd.DataFrame(bs_detail_rows)
    df_bs_unit   = pd.DataFrame(bs_unit_rows)

    bs_per_teknik = (df_bs_unit.groupby('Teknik')['BS_Konsistensi']
                     .agg(['mean','min','max','std']).round(4))
    bs_per_teknik.columns = ['Mean','Min','Max','Std']
    bs_per_teknik = bs_per_teknik.reindex(TEKNIK)

    print("\n  Skor Konsistensi BERTScore per Teknik:")
    print(f"  {'Teknik':15} {'Mean':8} {'Min':8} {'Max':8} {'Std':8}")
    for t in TEKNIK:
        r = bs_per_teknik.loc[t]
        print(f"  {t:15} {r['Mean']:.4f}   {r['Min']:.4f}   {r['Max']:.4f}   {r['Std']:.4f}")

    print("\n  Perbandingan Cosine vs BERTScore per Teknik:")
    print(f"  {'Teknik':15} {'Cosine':10} {'BERTScore':10} {'Selisih':10}")
    for t in TEKNIK:
        c = cosine_per_teknik.loc[t,'Mean']
        b = bs_per_teknik.loc[t,'Mean']
        print(f"  {t:15} {c:.4f}     {b:.4f}     {abs(c-b):.4f}")

except ImportError:
    print("  [SKIP] bert-score belum terinstall.")
    print("  Install: pip install bert-score torch")

#  6. Simpan ke Excel 
print("\n[6/6] Menyimpan hasil ke Excel...")

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

NAVY, BLUE, WHITE = "1F4E79", "2E75B6", "FFFFFF"
TC = {"Zero-Shot":"EBF3FB","Few-Shot":"FFF8E6","CoT":"F0FBF0","Role":"F5F0F8","Hybrid":"FFF0F0"}
TH = {"Zero-Shot":"1F4E79","Few-Shot":"7F4F00","CoT":"1E5628","Role":"5B0070","Hybrid":"7F0000"}

wb = Workbook()

# Sheet 1: Ringkasan per teknik
ws1 = wb.active
ws1.title = "Ringkasan Per Teknik"
ws1.sheet_view.showGridLines = False
n_cols = 9 if bs_available else 5
for i in range(1, n_cols+1):
    ws1.column_dimensions[get_column_letter(i)].width = 15

ws1.merge_cells(f"A1:{get_column_letter(n_cols)}1")
c = ws1["A1"]
c.value = ("RINGKASAN KONSISTENSI ANTAR RUN — "
           "Cosine Similarity (Formula Manual)" +
           (" + BERTScore (Simetris)" if bs_available else ""))
c.font = fnt(bold=True, size=11, color=WHITE)
c.fill = fill(NAVY); c.alignment = aln()

hdrs = ["Teknik","Cosine Mean","Cosine Min","Cosine Max","Cosine Std"]
if bs_available: hdrs += ["BS Mean","BS Min","BS Max","BS Std"]
for col, h in enumerate(hdrs, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY if col==1 else BLUE)
    c.alignment = aln(); c.border = bdr()

for i, teknik in enumerate(TEKNIK):
    row = i + 3
    bg = TC[teknik]; th = TH[teknik]
    vals = [teknik,
            cosine_per_teknik.loc[teknik,'Mean'],
            cosine_per_teknik.loc[teknik,'Min'],
            cosine_per_teknik.loc[teknik,'Max'],
            cosine_per_teknik.loc[teknik,'Std']]
    if bs_available:
        vals += [bs_per_teknik.loc[teknik,'Mean'],
                 bs_per_teknik.loc[teknik,'Min'],
                 bs_per_teknik.loc[teknik,'Max'],
                 bs_per_teknik.loc[teknik,'Std']]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = fnt(bold=(col==1), color=th if col==1 else "000000", size=10)
        c.fill = fill(bg); c.alignment = aln(h="left" if col==1 else "center")
        c.border = bdr()
        if col > 1: c.number_format = "0.0000"

# Sheet 2: Skor per unit (250 baris)
ws2 = wb.create_sheet("Skor Per Unit (250)")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"
hdrs2 = ["Pertanyaan","Sub-Bidang","Teknik",
          "Cosine Kons.","C.Min","C.Max","C.Std"]
if bs_available: hdrs2 += ["BS Kons.","BS.Min","BS.Max","BS.Std"]
for i, w in enumerate([14,24,14,12,10,10,10,12,10,10,10][:len(hdrs2)], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for col, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY if col<=3 else BLUE)
    c.alignment = aln(); c.border = bdr()

for i, rd in df_cosine_unit.iterrows():
    row = i + 2
    teknik = rd['Teknik']
    bg = TC.get(teknik,"FFFFFF"); th = TH.get(teknik,NAVY)
    vals = [rd['Pertanyaan'], rd['Sub_Bidang'], teknik,
            rd['Cosine_Konsistensi'], rd['Cosine_Min'],
            rd['Cosine_Max'], rd['Cosine_Std']]
    if bs_available:
        bs_row = df_bs_unit[(df_bs_unit['Pertanyaan']==rd['Pertanyaan']) &
                             (df_bs_unit['Teknik']==teknik)]
        if not bs_row.empty:
            br = bs_row.iloc[0]
            vals += [br['BS_Konsistensi'], br['BS_Min'], br['BS_Max'], br['BS_Std']]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = fnt(size=9, color=th if col==3 else "000000", bold=(col==3))
        c.fill = fill(bg)
        c.alignment = aln(h="left" if col<=2 else "center")
        c.border = bdr()
        if col > 3: c.number_format = "0.0000"

# Sheet 3: Cosine Detail (2500)
ws3 = wb.create_sheet("Cosine Detail (2500)")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A2"
for col, w in zip("ABCDEF", [14,24,14,8,8,12]):
    ws3.column_dimensions[col].width = w
for col, h in enumerate(["Pertanyaan","Sub-Bidang","Teknik","Run A","Run B","Cosine"], 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY if col<=3 else BLUE)
    c.alignment = aln(); c.border = bdr()
for i, rd in df_cosine_detail.iterrows():
    row = i + 2
    bg = TC.get(rd['Teknik'],"FFFFFF")
    for col, val in enumerate([rd['Pertanyaan'],rd['Sub_Bidang'],rd['Teknik'],
                                rd['Run_A'],rd['Run_B'],rd['Cosine']], 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = fnt(size=9); c.fill = fill(bg)
        c.alignment = aln(h="left" if col<=2 else "center")
        c.border = bdr()
        if col==6: c.number_format = "0.0000"

# Sheet 4: BERTScore Detail (2500)
if bs_available:
    ws4 = wb.create_sheet("BERTScore Detail (2500)")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "A2"
    for col, w in zip("ABCDEFGH", [14,24,14,8,8,12,12,14]):
        ws4.column_dimensions[col].width = w
    for col, h in enumerate(["Pertanyaan","Sub-Bidang","Teknik","Run A","Run B",
                               "BS F1(A→B)","BS F1(B→A)","BS Simetris"], 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = fnt(bold=True, color=WHITE, size=10)
        c.fill = fill(NAVY if col<=3 else BLUE)
        c.alignment = aln(); c.border = bdr()
    for i, rd in df_bs_detail.iterrows():
        row = i + 2
        bg = TC.get(rd['Teknik'],"FFFFFF")
        for col, val in enumerate([rd['Pertanyaan'],rd['Sub_Bidang'],rd['Teknik'],
                                    rd['Run_A'],rd['Run_B'],
                                    rd['BS_AB'],rd['BS_BA'],rd['BS_Sym']], 1):
            c = ws4.cell(row=row, column=col, value=val)
            c.font = fnt(size=9); c.fill = fill(bg)
            c.alignment = aln(h="left" if col<=2 else "center")
            c.border = bdr()
            if col>=6: c.number_format = "0.0000"

wb.save("cosine_bertscore_consistency.xlsx")
print("  Disimpan: cosine_bertscore_consistency.xlsx")

print("\n" + "=" * 70)
print("SELESAI!")
print("=" * 70)


# BAGIAN AKHIR: BOX PLOT COSINE DAN BERTSCORE

print("\n\n" + "=" * 70)
print("MEMBUAT BOX PLOT COSINE SIMILARITY DAN BERTSCORE")
print("=" * 70)

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

TEKNIK = ["Zero-Shot", "Few-Shot", "CoT", "Role", "Hybrid"]
TEKNIK_SHORT = ["ZS", "FS", "CoT", "Role", "Hybrid"]
TEKNIK_COLORS = {
    "Zero-Shot": "#4472C4",
    "Few-Shot":  "#ED7D31",
    "CoT":       "#70AD47",
    "Role":      "#7030A0",
    "Hybrid":    "#C00000",
}


df_cosine = pd.DataFrame(cosine_unit_rows)
df_cosine = df_cosine[df_cosine['Teknik'].isin(TEKNIK)]

#  Plot 1: Box Plot side-by-side Cosine vs BERTScore 
fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle(
    'Distribusi Skor Konsistensi Antar Run per Teknik Prompting\n'
    'Claude Sonnet 4.6 — 50 Pertanyaan × 5 Teknik (n=50 per teknik)',
    fontsize=13, fontweight='bold', y=1.02
)

# Subplot 1: Cosine Similarity 
cosine_data = [df_cosine[df_cosine['Teknik']==t]['Cosine_Konsistensi'].values
               for t in TEKNIK]
colors_list = [TEKNIK_COLORS[t] for t in TEKNIK]

bp1 = axes[0].boxplot(
    cosine_data,
    labels=TEKNIK_SHORT,
    patch_artist=True,
    medianprops=dict(color='white', linewidth=2.5),
    whiskerprops=dict(linewidth=1.5),
    capprops=dict(linewidth=2),
    flierprops=dict(marker='o', markersize=5, alpha=0.6),
    widths=0.6
)
for patch, color in zip(bp1['boxes'], colors_list):
    patch.set_facecolor(color)
    patch.set_alpha(0.8)
for flier, color in zip(bp1['fliers'], colors_list):
    flier.set(markerfacecolor=color, markeredgecolor=color)

axes[0].set_title('Cosine Similarity', fontsize=12, fontweight='bold', pad=10)
axes[0].set_ylabel('Skor Konsistensi', fontsize=11)
axes[0].set_ylim(0.45, 1.05)
axes[0].axhline(y=0.90, color='gray', linestyle='--', alpha=0.5, linewidth=1,
                label='Threshold 0,90 (sangat konsisten)')
axes[0].axhline(y=0.80, color='lightgray', linestyle=':', alpha=0.7, linewidth=1)
axes[0].grid(axis='y', alpha=0.3)
axes[0].legend(fontsize=9)

# Tambahkan nilai mean di atas setiap box
for i, (t, data) in enumerate(zip(TEKNIK, cosine_data)):
    mean_val = np.mean(data)
    axes[0].text(i+1, mean_val + 0.015, f'{mean_val:.3f}',
                ha='center', va='bottom', fontsize=9, fontweight='bold',
                color=TEKNIK_COLORS[t])

# Subplot 2: BERTScore 
if bs_available and not df_bs_unit.empty:
    df_bs = pd.DataFrame(bs_unit_rows)
    df_bs = df_bs[df_bs['Teknik'].isin(TEKNIK)]
    bs_data = [df_bs[df_bs['Teknik']==t]['BS_Konsistensi'].values for t in TEKNIK]

    bp2 = axes[1].boxplot(
        bs_data,
        labels=TEKNIK_SHORT,
        patch_artist=True,
        medianprops=dict(color='white', linewidth=2.5),
        whiskerprops=dict(linewidth=1.5),
        capprops=dict(linewidth=2),
        flierprops=dict(marker='o', markersize=5, alpha=0.6),
        widths=0.6
    )
    for patch, color in zip(bp2['boxes'], colors_list):
        patch.set_facecolor(color)
        patch.set_alpha(0.8)
    for flier, color in zip(bp2['fliers'], colors_list):
        flier.set(markerfacecolor=color, markeredgecolor=color)

    for i, (t, data) in enumerate(zip(TEKNIK, bs_data)):
        mean_val = np.mean(data)
        axes[1].text(i+1, mean_val + 0.008, f'{mean_val:.3f}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold',
                    color=TEKNIK_COLORS[t])
else:
    # Fallback: isi dengan data dummy dari ringkasan jika BERTScore tidak tersedia
    axes[1].text(0.5, 0.5, 'BERTScore tidak tersedia\n(jalankan dengan bert-score terinstall)',
                ha='center', va='center', transform=axes[1].transAxes,
                fontsize=11, color='gray')

axes[1].set_title('BERTScore (F1 Simetris)', fontsize=12, fontweight='bold', pad=10)
axes[1].set_ylabel('Skor Konsistensi', fontsize=11)
axes[1].set_ylim(0.65, 1.02)
axes[1].axhline(y=0.90, color='gray', linestyle='--', alpha=0.5, linewidth=1,
                label='Threshold 0,90')
axes[1].axhline(y=0.85, color='lightgray', linestyle=':', alpha=0.7, linewidth=1)
axes[1].grid(axis='y', alpha=0.3)
axes[1].legend(fontsize=9)

# Legenda warna teknik
legend_patches = [mpatches.Patch(facecolor=TEKNIK_COLORS[t], alpha=0.8, label=t)
                  for t in TEKNIK]
fig.legend(handles=legend_patches, loc='lower center', ncol=5,
           bbox_to_anchor=(0.5, -0.06), fontsize=10,
           title='Teknik Prompting', title_fontsize=10)

plt.tight_layout()
plt.savefig('boxplot_cosine_bertscore.png', dpi=150, bbox_inches='tight')
plt.close()
print("  Box plot disimpan: boxplot_cosine_bertscore.png")

# ── Plot 2: Box Plot gabungan (untuk perbandingan langsung) ───
if bs_available and not df_bs_unit.empty:
    fig2, ax = plt.subplots(figsize=(14, 6))

    positions_cosine = [1, 3, 5, 7, 9]
    positions_bs     = [2, 4, 6, 8, 10]

    bp_c = ax.boxplot(cosine_data, positions=positions_cosine, widths=0.7,
                      patch_artist=True,
                      medianprops=dict(color='white', linewidth=2),
                      whiskerprops=dict(linewidth=1.5),
                      capprops=dict(linewidth=1.5),
                      flierprops=dict(marker='o', markersize=4, alpha=0.5))
    for patch, color in zip(bp_c['boxes'], colors_list):
        patch.set_facecolor(color); patch.set_alpha(0.9)

    bp_b = ax.boxplot(bs_data, positions=positions_bs, widths=0.7,
                      patch_artist=True,
                      medianprops=dict(color='black', linewidth=2),
                      whiskerprops=dict(linewidth=1.5, linestyle='--'),
                      capprops=dict(linewidth=1.5),
                      flierprops=dict(marker='s', markersize=4, alpha=0.5))
    for patch, color in zip(bp_b['boxes'], colors_list):
        patch.set_facecolor(color); patch.set_alpha(0.4)
        patch.set_hatch('//')

    ax.set_xticks([1.5, 3.5, 5.5, 7.5, 9.5])
    ax.set_xticklabels(TEKNIK_SHORT, fontsize=11)
    ax.set_ylabel('Skor Konsistensi', fontsize=12)
    ax.set_title(
        'Perbandingan Distribusi Cosine Similarity vs BERTScore per Teknik Prompting\n'
        '(Box padat = Cosine Similarity | Box arsir = BERTScore)',
        fontsize=12, fontweight='bold'
    )
    ax.set_ylim(0.45, 1.05)
    ax.axhline(y=0.90, color='gray', linestyle='--', alpha=0.4, linewidth=1)
    ax.grid(axis='y', alpha=0.25)

    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='gray', alpha=0.9, label='Cosine Similarity (box padat)'),
        Patch(facecolor='gray', alpha=0.4, hatch='//', label='BERTScore (box arsir)'),
    ] + [Patch(facecolor=TEKNIK_COLORS[t], alpha=0.8, label=t) for t in TEKNIK]
    ax.legend(handles=legend_elements, loc='lower right', fontsize=9, ncol=2)

    plt.tight_layout()
    plt.savefig('boxplot_gabungan.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("  Box plot gabungan disimpan: boxplot_gabungan.png")

# Sheet Per Sub-Bidang + Sheet Perbandingan di Excel

print("\n[Tambahan] Menambahkan sheet ringkasan ke Excel...")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def xfill(h):  return PatternFill("solid", fgColor=h)
def xfont(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def xaln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def xbdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

NAVY, BLUE, WHITE = "1F4E79", "2E75B6", "FFFFFF"
TC = {"Zero-Shot":"EBF3FB","Few-Shot":"FFF8E6","CoT":"F0FBF0","Role":"F5F0F8","Hybrid":"FFF0F0"}
TH = {"Zero-Shot":"1F4E79","Few-Shot":"7F4F00","CoT":"1E5628","Role":"5B0070","Hybrid":"7F0000"}
SUB_MAP = {
    "AI dan Machine Learning":              "AI & ML",
    "Internet of Things":                   "IoT",
    "Jaringan Komputer dan Keamanan Siber": "Jarkom & Keamanan",
    "Komputasi Terdistribusi dan Cloud":    "Komputasi & Cloud",
    "Rekayasa Perangkat Lunak":             "RPL",
    "Sistem Informasi":                     "Sistem Informasi",
}

wb = load_workbook("cosine_bertscore_consistency.xlsx")

# ── Sheet: Per Sub-Bidang ────────────────────────────────────
if "Per Sub-Bidang (Laporan)" in wb.sheetnames:
    del wb["Per Sub-Bidang (Laporan)"]
ws_sub = wb.create_sheet("Per Sub-Bidang (Laporan)")
ws_sub.sheet_view.showGridLines = False

# Hitung per sub-bidang per teknik
df_c = pd.DataFrame(cosine_unit_rows)
df_b = pd.DataFrame(bs_unit_rows) if (bs_available and not df_bs_unit.empty) else pd.DataFrame()

sub_list = list(SUB_MAP.keys())

# Column widths
ws_sub.column_dimensions['A'].width = 22
for col in range(2, 12):
    ws_sub.column_dimensions[get_column_letter(col)].width = 9

# Row 1: Title
ws_sub.row_dimensions[1].height = 26
ws_sub.merge_cells("A1:K1")
c = ws_sub["A1"]
c.value = "TABEL: Skor Konsistensi Per Sub-Bidang (Cosine & BERTScore)"
c.font = xfont(bold=True, size=11, color=WHITE)
c.fill = xfill(NAVY); c.alignment = xaln()

# Row 2: Teknik header
ws_sub.row_dimensions[2].height = 22
ws_sub.cell(row=2, column=1, value="Sub-Bidang").font = xfont(bold=True, color=WHITE, size=10)
ws_sub.cell(row=2, column=1).fill = xfill(NAVY)
ws_sub.cell(row=2, column=1).alignment = xaln(h="left")
ws_sub.cell(row=2, column=1).border = xbdr()

for j, t in enumerate(TEKNIK):
    col1 = 2 + j*2
    ws_sub.merge_cells(start_row=2, start_column=col1, end_row=2, end_column=col1+1)
    c = ws_sub.cell(row=2, column=col1, value=t)
    c.font = xfont(bold=True, color=WHITE, size=9)
    c.fill = xfill(TH[t]); c.alignment = xaln(); c.border = xbdr()

# Row 3: Cos / BS sub-header
ws_sub.row_dimensions[3].height = 20
ws_sub.cell(row=3, column=1, value="").border = xbdr()
for j in range(5):
    col1 = 2 + j*2
    for off, lbl in enumerate(["Cos","BS"]):
        c = ws_sub.cell(row=3, column=col1+off, value=lbl)
        c.font = xfont(bold=True, color=WHITE, size=9)
        c.fill = xfill(BLUE); c.alignment = xaln(); c.border = xbdr()

# Data rows
for i, sub_full in enumerate(sub_list):
    row = i + 4
    ws_sub.row_dimensions[row].height = 22
    sub_short = SUB_MAP[sub_full]
    bg = "F5F9FF" if i%2==0 else "FFFFFF"

    c = ws_sub.cell(row=row, column=1, value=sub_short)
    c.font = xfont(size=10); c.fill = xfill(bg)
    c.alignment = xaln(h="left"); c.border = xbdr()

    best_cos = -1; best_bs = -1
    row_vals = []
    for t in TEKNIK:
        dc = df_c[(df_c['Sub_Bidang']==sub_full) & (df_c['Teknik']==t)]['Cosine_Konsistensi'].mean()
        if not df_b.empty:
            db = df_b[(df_b['Sub_Bidang']==sub_full) & (df_b['Teknik']==t)]['BS_Konsistensi'].mean()
        else:
            db = float('nan')
        row_vals.append((dc, db))
        if dc > best_cos: best_cos = dc
        if db > best_bs:  best_bs = db

    for j, (t, (dc, db)) in enumerate(zip(TEKNIK, row_vals)):
        col1 = 2 + j*2
        is_best_c = abs(dc - best_cos) < 0.0001
        is_best_b = (not pd.isna(db)) and abs(db - best_bs) < 0.0001

        c = ws_sub.cell(row=row, column=col1, value=round(dc, 4) if not pd.isna(dc) else "—")
        c.font = xfont(bold=is_best_c, size=10, color="375623" if is_best_c else "000000")
        c.fill = xfill("E2F0D9" if is_best_c else bg)
        c.alignment = xaln(); c.border = xbdr()
        c.number_format = "0.000"

        c = ws_sub.cell(row=row, column=col1+1, value=round(db, 4) if not pd.isna(db) else "—")
        c.font = xfont(bold=is_best_b, size=10, color="375623" if is_best_b else "000000")
        c.fill = xfill("E2F0D9" if is_best_b else bg)
        c.alignment = xaln(); c.border = xbdr()
        c.number_format = "0.000"

# ── Sheet: Perbandingan Dua Metode ───────────────────────────
if "Perbandingan Cos vs BS (Laporan)" in wb.sheetnames:
    del wb["Perbandingan Cos vs BS (Laporan)"]
ws_cmp = wb.create_sheet("Perbandingan Cos vs BS (Laporan)")
ws_cmp.sheet_view.showGridLines = False
for col, w in enumerate([20, 14, 16, 12, 36], 1):
    ws_cmp.column_dimensions[get_column_letter(col)].width = w

ws_cmp.row_dimensions[1].height = 26
ws_cmp.merge_cells("A1:E1")
c = ws_cmp["A1"]
c.value = "TABEL: Perbandingan Skor Konsistensi Cosine Similarity vs BERTScore"
c.font = xfont(bold=True, size=11, color=WHITE)
c.fill = xfill(NAVY); c.alignment = xaln()

ws_cmp.row_dimensions[2].height = 26
for col, h in enumerate(["Teknik","Cosine Mean","BERTScore Mean","|Selisih|","Interpretasi Divergensi"], 1):
    c = ws_cmp.cell(row=2, column=col, value=h)
    c.font = xfont(bold=True, color=WHITE, size=10)
    c.fill = xfill(NAVY if col==1 else BLUE)
    c.alignment = xaln(); c.border = xbdr()

# Hitung per teknik
cmp_rows = []
for t in TEKNIK:
    cm = df_c[df_c['Teknik']==t]['Cosine_Konsistensi'].mean()
    bm = df_b[df_b['Teknik']==t]['BS_Konsistensi'].mean() if not df_b.empty else float('nan')
    diff = abs(cm - bm) if not pd.isna(bm) else float('nan')
    cmp_rows.append({'Teknik': t, 'Cosine': cm, 'BS': bm, 'Diff': diff})

# Sort by Cosine descending
cmp_rows.sort(key=lambda x: x['Cosine'], reverse=True)

interp_map = {
    "Hybrid":    "Keduanya menunjukkan konsistensi yang tinggi baik pada level dokumen maupun level token.",
    "Few-Shot":  "Cosine similarity lebih tinggi karena format keluaran lebih konsisten, sedangkan variasi konten pada level token masih relatif lebih besar.",
    "Role":      "Sangat konsisten pada kedua metrik sehingga stabil baik pada level dokumen maupun token.",
    "CoT":       "BERTScore sedikit lebih tinggi, mengindikasikan bahwa penalaran bertahap membantu menjaga konsistensi konten pada level token.",
    "Zero-Shot": "Kedua metrik menunjukkan konsistensi paling rendah dengan variasi yang relatif lebih tinggi pada level dokumen maupun token.",
}

for i, row_data in enumerate(cmp_rows):
    row = i + 3
    ws_cmp.row_dimensions[row].height = 40
    t = row_data['Teknik']
    bg = TC.get(t, "F5F9FF")
    th = TH.get(t, NAVY)
    high_diff = (not pd.isna(row_data['Diff'])) and row_data['Diff'] >= 0.04

    vals = [t,
            round(row_data['Cosine'], 4),
            round(row_data['BS'], 4) if not pd.isna(row_data['BS']) else "—",
            round(row_data['Diff'], 4) if not pd.isna(row_data['Diff']) else "—",
            interp_map.get(t, "")]
    for col, val in enumerate(vals, 1):
        c = ws_cmp.cell(row=row, column=col, value=val)
        c.font = xfont(bold=(col==1), size=10,
                       color=th if col==1 else ("7F0000" if (col==4 and high_diff) else "000000"))
        c.fill = xfill(bg)
        c.alignment = xaln(h="left" if col in [1,5] else "center", wrap=True)
        c.border = xbdr()
        if col in [2,3,4]:
            c.number_format = "0.0000"

wb.save("cosine_bertscore_consistency.xlsx")

print("\nSelesai! File output:")
if bs_available:
    print("  boxplot_gabungan.png                (satu plot perbandingan)")