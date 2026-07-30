"""
==============================================================
ANALISIS BLEU + ROUGE + BERTSCORE — Konsistensi Antar Run
==============================================================
Bambang Istijab — 105222007 — Universitas Pertamina

REVISI METODOLOGI (sesuai arahan dosen pembimbing & penguji):
  - Cosine Similarity DIHAPUS sebagai metode pengukuran mandiri.
    Konsep cosine similarity tetap dibahas secara konseptual di
    sub-bab BERTScore, karena BERTScore secara internal
    menghitung cosine similarity antar token embedding.
  - Ditambahkan BLEU dan ROUGE sebagai metrik leksikal
    (kemiripan kata), melengkapi BERTScore yang mengukur
    kemiripan semantik (makna).

Tujuan Penelitian No. 3:
  Menganalisis konsistensi output kelima teknik prompting
  melalui mekanisme pengujian 5 run per pertanyaan per teknik.

Struktur perhitungan (SAMA untuk ketiga metrik, apple-to-apple):
  Untuk setiap unit (1 teknik x 1 pertanyaan):
    C(5,2) = 10 pasangan run
    Setiap pasangan dihitung DUA ARAH lalu dirata-ratakan
    (disimetriskan), karena BLEU/ROUGE/BERTScore secara alami
    asimetris (ada peran kandidat vs referensi).
    unit_score = mean(10 nilai simetris)
  Total: 250 unit x 10 pasangan = 2.500 perhitungan PER METRIK
  Grand total: 2.500 x 3 metrik = 7.500 perhitungan

==============================================================
"""

import pandas as pd
import numpy as np
from itertools import combinations
import warnings
warnings.filterwarnings('ignore')

import nltk
nltk.download('punkt', quiet=True)
nltk.download('punkt_tab', quiet=True)
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from rouge_score import rouge_scorer

print("=" * 70)
print("ANALISIS BLEU + ROUGE + BERTSCORE — Konsistensi Antar Run")
print("Apple-to-Apple: 250 unit x 10 pasangan = 2.500 nilai per metrik")
print("=" * 70)

# ── 1. Load data ──────────────────────────────────────────────
print("\n[1/7] Membaca Data_Eksperimen.xlsx...")
src = pd.read_excel("Data_Eksperimen.xlsx", sheet_name="Log Eksperimen", header=1)
src.columns = ['ID_Sesi', 'ID_Pertanyaan', 'Pertanyaan', 'Sub_Bidang',
               'Teknik', 'Run', 'Prompt', 'Output', 'Status']
src['Run'] = pd.to_numeric(src['Run'], errors='coerce').fillna(0).astype(int)
src['Output'] = src['Output'].astype(str)
src = src[src['Run'].isin([1, 2, 3, 4, 5])].reset_index(drop=True)

TEKNIK = ["Zero-Shot", "Few-Shot", "CoT", "Role", "Hybrid"]
pertanyaan_ids = sorted(src['ID_Pertanyaan'].unique())
sub_bidang_map = src.groupby('ID_Pertanyaan')['Sub_Bidang'].first().to_dict()
run_pairs = list(combinations([1, 2, 3, 4, 5], 2))  # 10 pasangan run

print(f"  Total sesi valid   : {len(src)}")
print(f"  Pertanyaan         : {len(pertanyaan_ids)}")
print(f"  Teknik             : {len(TEKNIK)}")
print(f"  Pasangan run       : {len(run_pairs)} (C(5,2))")
print(f"  Total unit         : {len(pertanyaan_ids) * len(TEKNIK)}")
print(f"  Total per metrik   : {len(pertanyaan_ids) * len(TEKNIK) * len(run_pairs)}")
print(f"  Grand total (3 metrik): {len(pertanyaan_ids) * len(TEKNIK) * len(run_pairs) * 3}")

# ── 2. Lookup dict ───────────────────────────────────────────
print("\n[2/7] Membangun lookup dictionary output...")
output_lookup = {}
for _, row in src.iterrows():
    key = (row['ID_Pertanyaan'], row['Teknik'], int(row['Run']))
    output_lookup[key] = str(row['Output'])
print(f"  Total entri lookup: {len(output_lookup)}")

# Sanity check
pid0 = pertanyaan_ids[0]
r1_sample = output_lookup.get((pid0, 'Zero-Shot', 1), "")
r2_sample = output_lookup.get((pid0, 'Zero-Shot', 2), "")
print(f"\n  Sanity check ({pid0}, Zero-Shot):")
print(f"  Run 1 ({len(r1_sample)} chars) vs Run 2 ({len(r2_sample)} chars)")
print(f"  Identik? {r1_sample == r2_sample} (harus False)")


# ══════════════════════════════════════════════════════════════
# FUNGSI METRIK
# ══════════════════════════════════════════════════════════════

smoothing = SmoothingFunction().method4

def hitung_bleu(text_ref: str, text_cand: str) -> float:
    """
    BLEU (Bilingual Evaluation Understudy) — Papineni, dkk. (2002)

    Mengukur precision n-gram: berapa proporsi n-gram pada teks
    kandidat yang juga muncul pada teks referensi, dengan
    brevity penalty untuk mencegah kandidat yang terlalu pendek.

    BLEU = BP x exp(Sum(w_n x log(p_n)))

    Menggunakan smoothing method4 (Chen & Cherry, 2014) karena
    teks panjang (multi-paragraf) berisiko menghasilkan precision
    n-gram tinggi = 0 (terutama pada n-gram besar), sehingga
    smoothing diperlukan agar skor tidak collapse ke nol.
    """
    ref_tokens = text_ref.split()
    cand_tokens = text_cand.split()
    if len(cand_tokens) == 0 or len(ref_tokens) == 0:
        return 0.0
    try:
        score = sentence_bleu(
            [ref_tokens], cand_tokens,
            weights=(0.25, 0.25, 0.25, 0.25),  # BLEU-4 standar
            smoothing_function=smoothing
        )
        return float(score)
    except Exception:
        return 0.0


rouge_scorer_obj = rouge_scorer.RougeScorer(['rouge1', 'rouge2', 'rougeL'], use_stemmer=False)

def hitung_rouge(text_ref: str, text_cand: str) -> dict:
    """
    ROUGE (Recall-Oriented Understudy for Gisting Evaluation) — Lin (2004)

    Mengukur recall n-gram: berapa proporsi n-gram pada teks
    referensi yang tercakup pada teks kandidat.

    ROUGE-1  : overlap unigram (kata tunggal)
    ROUGE-2  : overlap bigram (pasangan kata berurutan)
    ROUGE-L  : Longest Common Subsequence (LCS), menangkap
               kemiripan struktur kalimat secara lebih fleksibel

    F1 dari precision & recall digunakan sebagai skor akhir
    tiap varian, agar dapat dibandingkan apple-to-apple dengan
    BLEU dan BERTScore yang juga berbasis F1.
    """
    scores = rouge_scorer_obj.score(text_ref, text_cand)
    return {
        'rouge1': scores['rouge1'].fmeasure,
        'rouge2': scores['rouge2'].fmeasure,
        'rougeL': scores['rougeL'].fmeasure,
    }


print("\n[3/7] Verifikasi fungsi metrik pada sampel...")
bleu_test = hitung_bleu(r1_sample, r2_sample)
rouge_test = hitung_rouge(r1_sample, r2_sample)
print(f"  BLEU (run1 vs run2)    : {bleu_test:.4f}")
print(f"  ROUGE-1 (run1 vs run2) : {rouge_test['rouge1']:.4f}")
print(f"  ROUGE-2 (run1 vs run2) : {rouge_test['rouge2']:.4f}")
print(f"  ROUGE-L (run1 vs run2) : {rouge_test['rougeL']:.4f}")
print("  (Nilai < 1.0 mengonfirmasi kedua run memang berbeda teks)")


# ══════════════════════════════════════════════════════════════
# 4. HITUNG BLEU (2.500 pasangan, simetris)
# ══════════════════════════════════════════════════════════════
print("\n[4/7] Menghitung BLEU antar run (2.500 pasangan)...")

bleu_detail_rows = []
bleu_unit_rows = []
done = 0
total = len(pertanyaan_ids) * len(TEKNIK) * len(run_pairs)

for pid in pertanyaan_ids:
    sub = sub_bidang_map[pid]
    for teknik in TEKNIK:
        pair_scores = []
        for (r1, r2) in run_pairs:
            text_r1 = output_lookup.get((pid, teknik, r1), "")
            text_r2 = output_lookup.get((pid, teknik, r2), "")
            if not text_r1 or not text_r2:
                continue

            # BLEU_sym(i,j) = [BLEU(ref=r_i, cand=r_j) + BLEU(ref=r_j, cand=r_i)] / 2
            bleu_ab = hitung_bleu(text_r1, text_r2)  # r1=referensi, r2=kandidat
            bleu_ba = hitung_bleu(text_r2, text_r1)  # r2=referensi, r1=kandidat
            val_sym = (bleu_ab + bleu_ba) / 2

            pair_scores.append(val_sym)
            bleu_detail_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'Run_A': r1, 'Run_B': r2,
                'BLEU_AB': round(bleu_ab, 6), 'BLEU_BA': round(bleu_ba, 6),
                'BLEU_Sym': round(val_sym, 6),
            })
            done += 1
            if done % 250 == 0:
                print(f"  Progress BLEU: {done}/{total} ({done/total*100:.0f}%)")

        if pair_scores:
            bleu_unit_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'BLEU_Konsistensi': round(np.mean(pair_scores), 6),
                'BLEU_Min': round(min(pair_scores), 6),
                'BLEU_Max': round(max(pair_scores), 6),
                'BLEU_Std': round(np.std(pair_scores), 6),
            })

df_bleu_detail = pd.DataFrame(bleu_detail_rows)
df_bleu_unit = pd.DataFrame(bleu_unit_rows)
bleu_per_teknik = (df_bleu_unit.groupby('Teknik')['BLEU_Konsistensi']
                   .agg(['mean', 'min', 'max', 'std']).round(4).reindex(TEKNIK))
bleu_per_teknik.columns = ['Mean', 'Min', 'Max', 'Std']

print("\n  Skor Konsistensi BLEU per Teknik:")
print(bleu_per_teknik.to_string())


# ══════════════════════════════════════════════════════════════
# 5. HITUNG ROUGE (2.500 pasangan, simetris) — pakai ROUGE-L
# ══════════════════════════════════════════════════════════════
print("\n[5/7] Menghitung ROUGE antar run (2.500 pasangan)...")
print("  Menghitung ROUGE-1, ROUGE-2, ROUGE-L; ROUGE-L sebagai skor utama")

rouge_detail_rows = []
rouge_unit_rows = []
done = 0

for pid in pertanyaan_ids:
    sub = sub_bidang_map[pid]
    for teknik in TEKNIK:
        pair_scores_L = []
        pair_scores_1 = []
        pair_scores_2 = []
        for (r1, r2) in run_pairs:
            text_r1 = output_lookup.get((pid, teknik, r1), "")
            text_r2 = output_lookup.get((pid, teknik, r2), "")
            if not text_r1 or not text_r2:
                continue

            rg_ab = hitung_rouge(text_r1, text_r2)  # ref=r1, cand=r2
            rg_ba = hitung_rouge(text_r2, text_r1)  # ref=r2, cand=r1

            val_L = (rg_ab['rougeL'] + rg_ba['rougeL']) / 2
            val_1 = (rg_ab['rouge1'] + rg_ba['rouge1']) / 2
            val_2 = (rg_ab['rouge2'] + rg_ba['rouge2']) / 2

            pair_scores_L.append(val_L)
            pair_scores_1.append(val_1)
            pair_scores_2.append(val_2)

            rouge_detail_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'Run_A': r1, 'Run_B': r2,
                'ROUGE1_Sym': round(val_1, 6),
                'ROUGE2_Sym': round(val_2, 6),
                'ROUGEL_Sym': round(val_L, 6),
            })
            done += 1
            if done % 250 == 0:
                print(f"  Progress ROUGE: {done}/{total} ({done/total*100:.0f}%)")

        if pair_scores_L:
            rouge_unit_rows.append({
                'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                'ROUGE_Konsistensi': round(np.mean(pair_scores_L), 6),  # ROUGE-L sbg utama
                'ROUGE_Min': round(min(pair_scores_L), 6),
                'ROUGE_Max': round(max(pair_scores_L), 6),
                'ROUGE_Std': round(np.std(pair_scores_L), 6),
                'ROUGE1_Mean': round(np.mean(pair_scores_1), 6),
                'ROUGE2_Mean': round(np.mean(pair_scores_2), 6),
            })

df_rouge_detail = pd.DataFrame(rouge_detail_rows)
df_rouge_unit = pd.DataFrame(rouge_unit_rows)
rouge_per_teknik = (df_rouge_unit.groupby('Teknik')['ROUGE_Konsistensi']
                    .agg(['mean', 'min', 'max', 'std']).round(4).reindex(TEKNIK))
rouge_per_teknik.columns = ['Mean', 'Min', 'Max', 'Std']

print("\n  Skor Konsistensi ROUGE-L per Teknik:")
print(rouge_per_teknik.to_string())


# ══════════════════════════════════════════════════════════════
# 6. HITUNG BERTSCORE (2.500 pasangan, simetris)
# ══════════════════════════════════════════════════════════════
print("\n[6/7] Menghitung BERTScore antar run (2.500 pasangan)...")
bs_available = False
try:
    from bert_score import score as bs_score
    bs_available = True

    bs_detail_rows = []
    bs_unit_rows = []
    done_bs = 0

    for pid in pertanyaan_ids:
        sub = sub_bidang_map[pid]
        for teknik in TEKNIK:
            pair_scores = []
            for (r1, r2) in run_pairs:
                text_r1 = output_lookup.get((pid, teknik, r1), "")
                text_r2 = output_lookup.get((pid, teknik, r2), "")
                if not text_r1 or not text_r2:
                    continue

                _, _, F1_ab = bs_score([text_r1], [text_r2], lang="id", verbose=False)
                _, _, F1_ba = bs_score([text_r2], [text_r1], lang="id", verbose=False)
                val_sym = (F1_ab.item() + F1_ba.item()) / 2

                pair_scores.append(val_sym)
                bs_detail_rows.append({
                    'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                    'Run_A': r1, 'Run_B': r2,
                    'BS_AB': round(F1_ab.item(), 6), 'BS_BA': round(F1_ba.item(), 6),
                    'BS_Sym': round(val_sym, 6),
                })
                done_bs += 1
                if done_bs % 250 == 0:
                    print(f"  Progress BERTScore: {done_bs}/{total} ({done_bs/total*100:.0f}%)")

            if pair_scores:
                bs_unit_rows.append({
                    'Pertanyaan': pid, 'Sub_Bidang': sub, 'Teknik': teknik,
                    'BS_Konsistensi': round(np.mean(pair_scores), 6),
                    'BS_Min': round(min(pair_scores), 6),
                    'BS_Max': round(max(pair_scores), 6),
                    'BS_Std': round(np.std(pair_scores), 6),
                })

    df_bs_detail = pd.DataFrame(bs_detail_rows)
    df_bs_unit = pd.DataFrame(bs_unit_rows)
    bs_per_teknik = (df_bs_unit.groupby('Teknik')['BS_Konsistensi']
                     .agg(['mean', 'min', 'max', 'std']).round(4).reindex(TEKNIK))
    bs_per_teknik.columns = ['Mean', 'Min', 'Max', 'Std']

    print("\n  Skor Konsistensi BERTScore per Teknik:")
    print(bs_per_teknik.to_string())

except ImportError:
    print("  [SKIP] bert-score belum terinstall. Install: pip install bert-score torch")
    df_bs_detail = pd.DataFrame()
    df_bs_unit = pd.DataFrame()
    bs_per_teknik = pd.DataFrame()


# ══════════════════════════════════════════════════════════════
# 7. SIMPAN KE EXCEL + BOX PLOT
# ══════════════════════════════════════════════════════════════
print("\n[7/7] Menyimpan hasil ke Excel dan membuat box plot...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

NAVY, BLUE, WHITE = "1F4E79", "2E75B6", "FFFFFF"
TC = {"Zero-Shot": "EBF3FB", "Few-Shot": "FFF8E6", "CoT": "F0FBF0", "Role": "F5F0F8", "Hybrid": "FFF0F0"}
TH = {"Zero-Shot": "1F4E79", "Few-Shot": "7F4F00", "CoT": "1E5628", "Role": "5B0070", "Hybrid": "7F0000"}

wb = Workbook()

# ── Sheet 1: Ringkasan Per Teknik ─────────────────────────────
ws1 = wb.active
ws1.title = "Ringkasan Per Teknik"
ws1.sheet_view.showGridLines = False
n_cols = 13 if bs_available else 9
for i in range(1, n_cols + 1):
    ws1.column_dimensions[get_column_letter(i)].width = 14

ws1.merge_cells(f"A1:{get_column_letter(n_cols)}1")
c = ws1["A1"]
c.value = "RINGKASAN KONSISTENSI ANTAR RUN — BLEU + ROUGE-L + BERTScore"
c.font = fnt(bold=True, size=12, color=WHITE); c.fill = fill(NAVY); c.alignment = aln()

hdrs = ["Teknik", "BLEU Mean", "BLEU Min", "BLEU Max", "BLEU Std",
        "ROUGE-L Mean", "ROUGE-L Min", "ROUGE-L Max", "ROUGE-L Std"]
if bs_available:
    hdrs += ["BERTScore Mean", "BERTScore Min", "BERTScore Max", "BERTScore Std"]
for col, h in enumerate(hdrs, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=9)
    c.fill = fill(NAVY if col == 1 else BLUE); c.alignment = aln(); c.border = bdr()

for i, teknik in enumerate(TEKNIK):
    row = i + 3
    bg = TC[teknik]; th = TH[teknik]
    vals = [teknik,
            bleu_per_teknik.loc[teknik, 'Mean'], bleu_per_teknik.loc[teknik, 'Min'],
            bleu_per_teknik.loc[teknik, 'Max'], bleu_per_teknik.loc[teknik, 'Std'],
            rouge_per_teknik.loc[teknik, 'Mean'], rouge_per_teknik.loc[teknik, 'Min'],
            rouge_per_teknik.loc[teknik, 'Max'], rouge_per_teknik.loc[teknik, 'Std']]
    if bs_available:
        vals += [bs_per_teknik.loc[teknik, 'Mean'], bs_per_teknik.loc[teknik, 'Min'],
                  bs_per_teknik.loc[teknik, 'Max'], bs_per_teknik.loc[teknik, 'Std']]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = fnt(bold=(col == 1), color=th if col == 1 else "000000", size=10)
        c.fill = fill(bg); c.alignment = aln(h="left" if col == 1 else "center"); c.border = bdr()
        if col > 1:
            c.number_format = "0.0000"

# ── Sheet 2: Skor Per Unit (250) ───────────────────────────────
ws2 = wb.create_sheet("Skor Per Unit (250)")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"
hdrs2 = ["Pertanyaan", "Sub-Bidang", "Teknik", "BLEU Kons.", "ROUGE-L Kons.", "ROUGE-1", "ROUGE-2"]
if bs_available:
    hdrs2 += ["BERTScore Kons."]
for i, w in enumerate([14, 24, 14] + [13] * (len(hdrs2) - 3), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for col, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=9)
    c.fill = fill(NAVY if col <= 3 else BLUE); c.alignment = aln(); c.border = bdr()

for i, rd in df_bleu_unit.iterrows():
    row = i + 2
    teknik = rd['Teknik']; bg = TC.get(teknik, "FFFFFF"); th = TH.get(teknik, NAVY)
    rouge_row = df_rouge_unit[(df_rouge_unit['Pertanyaan'] == rd['Pertanyaan']) & (df_rouge_unit['Teknik'] == teknik)]
    rouge_row = rouge_row.iloc[0] if not rouge_row.empty else None
    vals = [rd['Pertanyaan'], rd['Sub_Bidang'], teknik, rd['BLEU_Konsistensi'],
            rouge_row['ROUGE_Konsistensi'] if rouge_row is not None else "",
            rouge_row['ROUGE1_Mean'] if rouge_row is not None else "",
            rouge_row['ROUGE2_Mean'] if rouge_row is not None else ""]
    if bs_available:
        bs_row = df_bs_unit[(df_bs_unit['Pertanyaan'] == rd['Pertanyaan']) & (df_bs_unit['Teknik'] == teknik)]
        vals.append(bs_row.iloc[0]['BS_Konsistensi'] if not bs_row.empty else "")
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = fnt(size=9, color=th if col == 3 else "000000", bold=(col == 3))
        c.fill = fill(bg); c.alignment = aln(h="left" if col <= 2 else "center"); c.border = bdr()
        if col > 3:
            c.number_format = "0.0000"

# ── Sheet 3: Detail BLEU (2500) ────────────────────────────────
ws3 = wb.create_sheet("BLEU Detail (2500)")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A2"
for col, w in zip("ABCDEFG", [14, 24, 14, 8, 8, 12, 12]):
    ws3.column_dimensions[col].width = w
for col, h in enumerate(["Pertanyaan", "Sub-Bidang", "Teknik", "Run A", "Run B", "BLEU(A→B)", "BLEU Simetris"], 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=9)
    c.fill = fill(NAVY if col <= 3 else BLUE); c.alignment = aln(); c.border = bdr()
for i, rd in df_bleu_detail.iterrows():
    row = i + 2
    bg = TC.get(rd['Teknik'], "FFFFFF")
    vals = [rd['Pertanyaan'], rd['Sub_Bidang'], rd['Teknik'], rd['Run_A'], rd['Run_B'], rd['BLEU_AB'], rd['BLEU_Sym']]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = fnt(size=9); c.fill = fill(bg)
        c.alignment = aln(h="left" if col <= 2 else "center"); c.border = bdr()
        if col >= 6:
            c.number_format = "0.0000"

# ── Sheet 4: Detail ROUGE (2500) ───────────────────────────────
ws4 = wb.create_sheet("ROUGE Detail (2500)")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A2"
for col, w in zip("ABCDEFGH", [14, 24, 14, 8, 8, 12, 12, 12]):
    ws4.column_dimensions[col].width = w
for col, h in enumerate(["Pertanyaan", "Sub-Bidang", "Teknik", "Run A", "Run B", "ROUGE-1", "ROUGE-2", "ROUGE-L"], 1):
    c = ws4.cell(row=1, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=9)
    c.fill = fill(NAVY if col <= 3 else BLUE); c.alignment = aln(); c.border = bdr()
for i, rd in df_rouge_detail.iterrows():
    row = i + 2
    bg = TC.get(rd['Teknik'], "FFFFFF")
    vals = [rd['Pertanyaan'], rd['Sub_Bidang'], rd['Teknik'], rd['Run_A'], rd['Run_B'],
            rd['ROUGE1_Sym'], rd['ROUGE2_Sym'], rd['ROUGEL_Sym']]
    for col, val in enumerate(vals, 1):
        c = ws4.cell(row=row, column=col, value=val)
        c.font = fnt(size=9); c.fill = fill(bg)
        c.alignment = aln(h="left" if col <= 2 else "center"); c.border = bdr()
        if col >= 6:
            c.number_format = "0.0000"

# ── Sheet 5: Detail BERTScore (2500) ───────────────────────────
if bs_available:
    ws5 = wb.create_sheet("BERTScore Detail (2500)")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = "A2"
    for col, w in zip("ABCDEFGH", [14, 24, 14, 8, 8, 12, 12, 12]):
        ws5.column_dimensions[col].width = w
    for col, h in enumerate(["Pertanyaan", "Sub-Bidang", "Teknik", "Run A", "Run B", "BS(A→B)", "BS(B→A)", "BS Simetris"], 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = fnt(bold=True, color=WHITE, size=9)
        c.fill = fill(NAVY if col <= 3 else BLUE); c.alignment = aln(); c.border = bdr()
    for i, rd in df_bs_detail.iterrows():
        row = i + 2
        bg = TC.get(rd['Teknik'], "FFFFFF")
        vals = [rd['Pertanyaan'], rd['Sub_Bidang'], rd['Teknik'], rd['Run_A'], rd['Run_B'],
                rd['BS_AB'], rd['BS_BA'], rd['BS_Sym']]
        for col, val in enumerate(vals, 1):
            c = ws5.cell(row=row, column=col, value=val)
            c.font = fnt(size=9); c.fill = fill(bg)
            c.alignment = aln(h="left" if col <= 2 else "center"); c.border = bdr()
            if col >= 6:
                c.number_format = "0.0000"

# ── Sheet 6: Per Sub-Bidang ─────────────────────────────────────
ws6 = wb.create_sheet("Per Sub-Bidang")
ws6.sheet_view.showGridLines = False
SUB_MAP = {
    "AI dan Machine Learning": "AI & ML", "Internet of Things": "IoT",
    "Jaringan Komputer dan Keamanan Siber": "Jarkom & Keamanan",
    "Komputasi Terdistribusi dan Cloud": "Komputasi & Cloud",
    "Rekayasa Perangkat Lunak": "RPL", "Sistem Informasi": "Sistem Informasi",
}
sub_list = list(SUB_MAP.keys())
n_metric = 3 if bs_available else 2
ws6.column_dimensions['A'].width = 20
for col in range(2, 2 + 5 * n_metric):
    ws6.column_dimensions[get_column_letter(col)].width = 9

ws6.merge_cells(f"A1:{get_column_letter(1 + 5*n_metric)}1")
c = ws6["A1"]
c.value = "SKOR KONSISTENSI PER SUB-BIDANG (BLEU / ROUGE-L / BERTScore)"
c.font = fnt(bold=True, size=11, color=WHITE); c.fill = fill(NAVY); c.alignment = aln()

col_i = 2
for t in TEKNIK:
    span = n_metric
    ws6.merge_cells(start_row=2, start_column=col_i, end_row=2, end_column=col_i + span - 1)
    c = ws6.cell(row=2, column=col_i, value=t)
    c.font = fnt(bold=True, color=WHITE, size=9); c.fill = fill(TH[t]); c.alignment = aln(); c.border = bdr()
    labels = ["BL", "RG"] + (["BS"] if bs_available else [])
    for j, lbl in enumerate(labels):
        c = ws6.cell(row=3, column=col_i + j, value=lbl)
        c.font = fnt(bold=True, color=WHITE, size=8); c.fill = fill(BLUE); c.alignment = aln(); c.border = bdr()
    col_i += span

ws6.cell(row=2, column=1, value="Sub-Bidang").font = fnt(bold=True, color=WHITE, size=9)
ws6.cell(row=2, column=1).fill = fill(NAVY); ws6.cell(row=2, column=1).alignment = aln(h="left")
ws6.merge_cells("A2:A3")

for i, sub_full in enumerate(sub_list):
    row = i + 4
    bg = "F5F9FF" if i % 2 == 0 else "FFFFFF"
    c = ws6.cell(row=row, column=1, value=SUB_MAP[sub_full])
    c.font = fnt(size=9); c.fill = fill(bg); c.alignment = aln(h="left"); c.border = bdr()
    col_i = 2
    for t in TEKNIK:
        bl = df_bleu_unit[(df_bleu_unit['Sub_Bidang'] == sub_full) & (df_bleu_unit['Teknik'] == t)]['BLEU_Konsistensi'].mean()
        rg = df_rouge_unit[(df_rouge_unit['Sub_Bidang'] == sub_full) & (df_rouge_unit['Teknik'] == t)]['ROUGE_Konsistensi'].mean()
        vals = [bl, rg]
        if bs_available:
            bsv = df_bs_unit[(df_bs_unit['Sub_Bidang'] == sub_full) & (df_bs_unit['Teknik'] == t)]['BS_Konsistensi'].mean()
            vals.append(bsv)
        for j, v in enumerate(vals):
            c = ws6.cell(row=row, column=col_i + j, value=round(v, 4) if not pd.isna(v) else "-")
            c.font = fnt(size=8); c.fill = fill(bg); c.alignment = aln(); c.border = bdr()
            c.number_format = "0.000"
        col_i += n_metric

wb.save("bleu_rouge_bertscore_consistency.xlsx")
print("  Disimpan: bleu_rouge_bertscore_consistency.xlsx")

# ── Box plot 3 metrik ────────────────────────────────────────
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

TEKNIK_SHORT = ["ZS", "FS", "CoT", "Role", "Hybrid"]
TEKNIK_COLORS = {"Zero-Shot": "#4472C4", "Few-Shot": "#ED7D31", "CoT": "#70AD47",
                  "Role": "#7030A0", "Hybrid": "#C00000"}
colors_list = [TEKNIK_COLORS[t] for t in TEKNIK]

bleu_data = [df_bleu_unit[df_bleu_unit['Teknik'] == t]['BLEU_Konsistensi'].values for t in TEKNIK]
rouge_data = [df_rouge_unit[df_rouge_unit['Teknik'] == t]['ROUGE_Konsistensi'].values for t in TEKNIK]

n_plots = 3 if bs_available else 2
fig, axes = plt.subplots(1, n_plots, figsize=(6 * n_plots, 6))
fig.suptitle('Distribusi Skor Konsistensi Antar Run per Teknik Prompting\n'
             'Claude Sonnet 4.6 — 50 Pertanyaan x 5 Teknik (n=50 per teknik)',
             fontsize=13, fontweight='bold', y=1.02)

def plot_box(ax, data, title, ylim):
    bp = ax.boxplot(data, labels=TEKNIK_SHORT, patch_artist=True,
                     medianprops=dict(color='white', linewidth=2.5),
                     whiskerprops=dict(linewidth=1.5), capprops=dict(linewidth=2),
                     flierprops=dict(marker='o', markersize=5, alpha=0.6), widths=0.6)
    for patch, color in zip(bp['boxes'], colors_list):
        patch.set_facecolor(color); patch.set_alpha(0.8)
    for flier, color in zip(bp['fliers'], colors_list):
        flier.set(markerfacecolor=color, markeredgecolor=color)
    ax.set_title(title, fontsize=12, fontweight='bold', pad=10)
    ax.set_ylabel('Skor Konsistensi', fontsize=11)
    ax.set_ylim(*ylim)
    ax.grid(axis='y', alpha=0.3)
    for i, (t, d) in enumerate(zip(TEKNIK, data)):
        if len(d) > 0:
            mean_val = np.mean(d)
            ax.text(i + 1, mean_val + (ylim[1]-ylim[0])*0.02, f'{mean_val:.3f}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold', color=TEKNIK_COLORS[t])

plot_box(axes[0], bleu_data, 'BLEU', (0, max(0.6, max(max(d) for d in bleu_data if len(d)>0)*1.15)))
plot_box(axes[1], rouge_data, 'ROUGE-L', (0, max(0.7, max(max(d) for d in rouge_data if len(d)>0)*1.15)))

if bs_available:
    bs_data = [df_bs_unit[df_bs_unit['Teknik'] == t]['BS_Konsistensi'].values for t in TEKNIK]
    plot_box(axes[2], bs_data, 'BERTScore', (0.6, 1.02))

legend_patches = [Patch(facecolor=TEKNIK_COLORS[t], alpha=0.8, label=t) for t in TEKNIK]
fig.legend(handles=legend_patches, loc='lower center', ncol=5,
           bbox_to_anchor=(0.5, -0.05), fontsize=10, title='Teknik Prompting', title_fontsize=10)

plt.tight_layout()
plt.savefig('boxplot_bleu_rouge_bertscore.png', dpi=150, bbox_inches='tight')
plt.close()
print("  Box plot disimpan: boxplot_bleu_rouge_bertscore.png")

print("\n" + "=" * 70)
print("SELESAI!")
print("=" * 70)
print(f"""
Ringkasan akhir (rata-rata seluruh teknik):
  BLEU      : {df_bleu_unit['BLEU_Konsistensi'].mean():.4f}
  ROUGE-L   : {df_rouge_unit['ROUGE_Konsistensi'].mean():.4f}
  {"BERTScore : " + f"{df_bs_unit['BS_Konsistensi'].mean():.4f}" if bs_available else ""}

File output:
  bleu_rouge_bertscore_consistency.xlsx  (6 sheet)
  boxplot_bleu_rouge_bertscore.png
""")
